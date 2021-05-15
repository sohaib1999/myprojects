import openpyxl  as xl

from openpyxl.chart import BarChart, Reference

wb =xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']

# prints particular cell value
def printCell(row,column):
  cell=sheet.cell(row,column)
  print(cell.value)

# automaticaly writes over excel file
def automate():
    i=2
    while i <= sheet.max_row:
        cell1=sheet.cell(i,4)
        cell2 = sheet.cell(i, 3)
        cell1.value=cell2.value*0.7
        i=i+1

# adds Barchart automaticaly
def addChart():
    values =Reference(sheet, min_row=2, max_row=4, min_col=4, max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'a5')

printCell(3,2)

automate()

addChart()

wb.save('transactions.xlsx')